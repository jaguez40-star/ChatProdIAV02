"""Construcción de hojas de Excel en memoria para los tests de extractores.

**Por qué existen dos clases de test para lo mismo.** Los tests contra el `.xlsm` real
verifican *fidelidad*: que el extractor saque de verdad los datos de la hoja pactada. Pero
llevan el marcador `muestras` y no corren en CI, porque esos archivos viven fuera del repo.
Sin nada más, el código de los 17 extractores contaría como no cubierto y hundiría el
umbral de cobertura conforme se fueran portando (hallazgo H5 del plan).

Estas hojas sintéticas cubren ese hueco: replican **solo el andamiaje mínimo** que cada
extractor busca (la fila de encabezado, la columna de anclaje, un par de filas de datos) y
ejercitan sus ramas dentro de CI, sin depender de ningún archivo externo.

Ninguna de las dos sustituye a la otra: la sintética prueba que la lógica funciona, la real
prueba que las posiciones son las correctas.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def hoja_desde_celdas(celdas: dict[tuple[int, int], Any]) -> Worksheet:
    """Hoja con los valores indicados en `{(fila, columna): valor}` (1-based, como Excel)."""
    libro = Workbook()
    hoja = libro.active
    assert hoja is not None
    for (fila, columna), valor in celdas.items():
        hoja.cell(row=fila, column=columna, value=valor)
    return hoja


def hoja_desde_filas(filas: list[list[Any]]) -> Worksheet:
    """Hoja construida fila a fila — más legible cuando el layout es secuencial."""
    libro = Workbook()
    hoja = libro.active
    assert hoja is not None
    for indice, valores in enumerate(filas, start=1):
        for columna, valor in enumerate(valores, start=1):
            if valor is not None:
                hoja.cell(row=indice, column=columna, value=valor)
    return hoja
