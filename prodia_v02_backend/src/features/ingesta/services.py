"""Orquestador del ETL de Ingesta.

Coordina el flujo completo de un `.xlsm`: detectar su tipo, aterrizar las hojas en
`bronze.*`, pasar las modeladas por sus 17 extractores y volcarlas a `core.fact_tabla_hoja`.

## La transacción y lo que promete

Todo ocurre dentro de **una sola transacción** (la que abre `get_prod_tx`). Es lo que
hace la ingesta atómica: si falla la hoja 30 de 37, no queda un reporte a medio cargar
que nadie sabría distinguir de uno completo.

El precio es que la transacción vive minutos y retiene locks. Por eso lo primero que hace
el servicio es tomar un `pg_advisory_xact_lock` por fecha de reporte: dos ingestas de la
misma fecha se serializan de forma explícita, en vez de bloquearse a ciegas dentro de
PostgreSQL o entrar en deadlock al tocar las dimensiones en distinto orden (G4).

## Por qué los eventos dicen "procesada" y no "ok"

El origen emitía `ok` por hoja **dentro** de la transacción, así que un fallo posterior
revertía todo mientras el usuario ya había visto verde (G2). Aquí las hojas se reportan
como `procesada` —insertada, pendiente de confirmar— y solo el `EventoFin` con estado
`confirmado` afirma que los datos quedaron guardados.

El servicio **no hace commit**: eso es responsabilidad de la dependencia `get_prod_tx`.
Así, quien decide confirmar es quien abrió la transacción, y el orquestador no puede
confirmar a medias por error.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Callable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.core.logger import get_logger
from src.features.ingesta.detector import HOJAS_RAW, tiene_raw
from src.features.ingesta.extractores import extractores_aplicables
from src.features.ingesta.repositories import IngestaRepository
from src.features.ingesta.schemas import (
    EventoHoja,
    EventoIngesta,
    EventoInicio,
    HojaIngerida,
    ResultadoIngesta,
    TablaIngerida,
)
from src.features.ingesta.transforms import BZ_DIA, BZ_MES, BZ_PRG

logger = get_logger("ingesta.services")

# Hoja RAW → (tabla bronze, columnas). El orden de columnas es significativo.
DESTINOS_BRONZE: list[tuple[str, str, list[str]]] = [
    ("BDP_datos_dia", "bdp_datos_dia", BZ_DIA),
    ("BDP_datos_mes", "bdp_datos_mes", BZ_MES),
    ("BDP_Programa", "bdp_programa", BZ_PRG),
]

# Un observador recibe cada evento; si no hay, la ingesta corre igual.
Observador = Callable[[EventoIngesta], None]


class IngestaService:
    """Ejecuta la ingesta de un archivo. No abre ni confirma la transacción."""

    def __init__(
        self, repositorio: IngestaRepository, observador: Observador | None = None
    ) -> None:
        self._repo = repositorio
        self._observador = observador

    def _emitir(self, evento: EventoIngesta) -> None:
        """Publica un evento. Un observador que falle NUNCA tumba la ingesta: el
        progreso es informativo y el dato es lo que importa."""
        if self._observador is None:
            return
        try:
            self._observador(evento)
        except Exception:  # noqa: BLE001 — el progreso no puede romper el ETL
            logger.warning("observador_de_progreso_fallo", exc_info=True)

    def ingerir(self, ruta: Path) -> ResultadoIngesta:
        """Procesa el archivo entero y devuelve el resumen.

        Cualquier excepción se propaga a propósito: `get_prod_tx` la convierte en
        rollback, y el router la traduce a un `EventoFin` con estado `revertido`.
        """
        libro = load_workbook(ruta, read_only=True, data_only=True, keep_links=False)
        try:
            return self._ingerir_libro(libro, ruta)
        finally:
            libro.close()

    def _ingerir_libro(self, libro: Any, ruta: Path) -> ResultadoIngesta:
        nombres_de_hoja: list[str] = list(libro.sheetnames)
        trae_raw = tiene_raw(set(nombres_de_hoja))
        tipo_archivo = "NEW" if trae_raw else "STD"

        self._emitir(
            EventoInicio(
                archivo=ruta.name,
                tipo_archivo=tipo_archivo,
                hojas=nombres_de_hoja,
                total=len(nombres_de_hoja),
            )
        )

        reporte_id, fecha_reporte = self._repo.upsert_reporte(ruta, trae_raw)
        # Antes de escribir nada más: serializa contra otra ingesta de la misma fecha.
        self._repo.tomar_bloqueo_de_reporte(fecha_reporte)

        filas_por_destino: dict[str, int] = {}
        hojas: list[HojaIngerida] = []
        tablas_vacias: list[str] = []

        if trae_raw:
            self._aterrizar_hojas_raw(libro, reporte_id, filas_por_destino, hojas)
        self._aterrizar_resto_de_hojas(
            libro, nombres_de_hoja, reporte_id, filas_por_destino, hojas
        )
        self._volcar_hojas_modeladas(
            libro, nombres_de_hoja, reporte_id, filas_por_destino, hojas, tablas_vacias
        )

        resultado = ResultadoIngesta(
            archivo=ruta.name,
            reporte_id=reporte_id,
            fecha_reporte=fecha_reporte,
            tipo_archivo=tipo_archivo,
            tiene_raw=trae_raw,
            filas_por_destino=filas_por_destino,
            hojas=hojas,
            tablas_vacias=tablas_vacias,
        )
        logger.info(
            "ingesta_procesada",
            archivo=ruta.name,
            reporte_id=reporte_id,
            tipo=tipo_archivo,
            filas=resultado.total_filas,
            tablas_vacias=len(tablas_vacias),
        )
        return resultado

    # ── Capa bronze ─────────────────────────────────────────────────────────

    def _aterrizar_hojas_raw(
        self,
        libro: Any,
        reporte_id: int,
        filas_por_destino: dict[str, int],
        hojas: list[HojaIngerida],
    ) -> None:
        """Las tres hojas planas van a sus tablas bronze tipadas."""
        for nombre_hoja, tabla, columnas in DESTINOS_BRONZE:
            self._emitir(EventoHoja(hoja=nombre_hoja, estado="procesando"))
            filas_hoja = list(libro[nombre_hoja].iter_rows(values_only=True))
            # Se salta el encabezado: bronze guarda datos, no la fila de títulos.
            total = self._repo.aterrizar_hoja_tipada(
                tabla, columnas, reporte_id, filas_hoja[1:]
            )
            destino = f"bronze.{tabla}"
            self._repo.registrar_en_bitacora(
                reporte_id, nombre_hoja, destino, total, total
            )
            filas_por_destino[destino] = total
            hojas.append(HojaIngerida(hoja=nombre_hoja, destino=destino, filas=total))
            self._emitir(
                EventoHoja(
                    hoja=nombre_hoja,
                    estado="procesada" if total else "vacia",
                    destino=destino,
                    filas=total,
                )
            )

    def _aterrizar_resto_de_hojas(
        self,
        libro: Any,
        nombres_de_hoja: list[str],
        reporte_id: int,
        filas_por_destino: dict[str, int],
        hojas: list[HojaIngerida],
    ) -> None:
        """Todas las demás hojas se preservan como JSONB en `bronze.hoja_landing`.

        Se aterriza **cualquier** hoja, esté modelada o no: bronze es la copia fiel del
        archivo, y una hoja que hoy nadie modela puede necesitarse mañana.
        """
        total_hojas = 0
        for nombre_hoja in nombres_de_hoja:
            if nombre_hoja in HOJAS_RAW:
                continue  # ya fueron a su tabla tipada
            self._emitir(EventoHoja(hoja=nombre_hoja, estado="procesando"))
            filas_hoja = list(libro[nombre_hoja].iter_rows(values_only=True))
            total = self._repo.aterrizar_hoja_generica(
                nombre_hoja, reporte_id, filas_hoja
            )
            self._repo.registrar_en_bitacora(
                reporte_id, nombre_hoja, "bronze.hoja_landing", total, total
            )
            total_hojas += 1
            self._emitir(
                EventoHoja(
                    hoja=nombre_hoja,
                    estado="procesada" if total else "vacia",
                    destino="bronze.hoja_landing",
                    filas=total,
                )
            )
        filas_por_destino["bronze.hoja_landing"] = total_hojas

    # ── Hojas modeladas → core.fact_tabla_hoja ──────────────────────────────

    def _volcar_hojas_modeladas(
        self,
        libro: Any,
        nombres_de_hoja: list[str],
        reporte_id: int,
        filas_por_destino: dict[str, int],
        hojas: list[HojaIngerida],
        tablas_vacias: list[str],
    ) -> None:
        """Pasa cada hoja modelada por su extractor y vuelca el resultado.

        Un extractor que revienta **no aborta la ingesta entera**: se reporta como
        `error` en esa hoja y el resto continúa. Es deliberado — perder una hoja por un
        cambio de layout no debería costar las otras dieciséis. El fallo sí queda
        registrado en la bitácora y en el log.
        """
        for nombre_hoja, extractor in extractores_aplicables(nombres_de_hoja):
            self._emitir(EventoHoja(hoja=nombre_hoja, estado="procesando"))
            try:
                resultado = extractor(libro[nombre_hoja])
            except Exception as excepcion:  # noqa: BLE001 — una hoja no tumba el resto
                logger.error(
                    "extractor_fallo",
                    hoja=nombre_hoja,
                    error=str(excepcion),
                    exc_info=True,
                )
                self._repo.registrar_en_bitacora(
                    reporte_id,
                    nombre_hoja,
                    "core.fact_tabla_hoja",
                    0,
                    0,
                    estado="ERROR",
                    mensaje=str(excepcion)[:500],
                )
                self._emitir(
                    EventoHoja(
                        hoja=nombre_hoja,
                        estado="error",
                        destino="core.fact_tabla_hoja",
                        detalle=str(excepcion)[:300],
                    )
                )
                continue

            insertadas = self._repo.reemplazar_tablas_de_hoja(
                reporte_id, nombre_hoja, resultado.filas
            )
            conteo = self._contar_por_tabla(insertadas, resultado.tablas_declaradas)
            for tabla in conteo:
                if tabla.filas == 0:
                    tablas_vacias.append(f"{nombre_hoja} → {tabla.tabla_label}")

            total = len(insertadas)
            filas_por_destino[f"fact_tabla_hoja::{nombre_hoja}"] = total
            self._repo.registrar_en_bitacora(
                reporte_id, nombre_hoja, "core.fact_tabla_hoja", total, total
            )
            hojas.append(
                HojaIngerida(
                    hoja=nombre_hoja,
                    destino="core.fact_tabla_hoja",
                    filas=total,
                    tablas=conteo,
                )
            )
            self._emitir(
                EventoHoja(
                    hoja=nombre_hoja,
                    estado="procesada" if total else "vacia",
                    destino="core.fact_tabla_hoja",
                    filas=total,
                    tablas=conteo,
                )
            )

    @staticmethod
    def _contar_por_tabla(
        insertadas: list[Any], declaradas: list[tuple[int, str]]
    ) -> list[TablaIngerida]:
        """Filas por tabla lógica, **incluyendo las declaradas que salieron en cero**.

        Esa inclusión es el corazón de G5: una tabla vacía tiene que aparecer en la lista
        para que se vea que existe y no produjo nada. Si solo se contaran las que tienen
        filas, un cambio de layout se manifestaría como una tabla que simplemente
        desapareció del visor, sin que nadie lo notase.
        """
        conteo: dict[tuple[int, str], int] = {}
        for fila in insertadas:
            clave = (fila.tabla_idx, fila.tabla_label)
            conteo[clave] = conteo.get(clave, 0) + 1

        if declaradas:
            return [
                TablaIngerida(
                    tabla_idx=indice,
                    tabla_label=etiqueta,
                    filas=conteo.get((indice, etiqueta), 0),
                )
                for indice, etiqueta in declaradas
            ]
        return [
            TablaIngerida(tabla_idx=indice, tabla_label=etiqueta, filas=total)
            for (indice, etiqueta), total in sorted(conteo.items())
        ]


def fecha_de_reporte_valida(nombre_archivo: str) -> dt.date | None:
    """Fecha embebida en el nombre del archivo, o `None` si no la trae.

    Se expone aparte del repositorio para que la API pueda rechazar un archivo sin fecha
    **antes** de abrir la transacción y empezar a escribir.
    """
    from src.features.ingesta.repositories import _FECHA_EN_NOMBRE

    coincidencia = _FECHA_EN_NOMBRE.search(nombre_archivo)
    if coincidencia is None:
        return None
    from src.features.ingesta.celdas import to_date

    return to_date(coincidencia.group(1))
