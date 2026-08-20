"""Extractores de las hojas de detalle — portados literalmente del sistema viejo.

Cinco hojas de grano fino, muy distintas de las modeladas:

- `TD_datos_dia` y `DATOS_MES` son **tablas dinámicas** de Excel: cabeceras jerárquicas
  y etiquetas de fila que solo aparecen en la primera fila de cada grupo.
- `BDP_datos_dia`, `BDP_datos_mes` y `BDP_Programa` son **tablas planas**: una cabecera
  en la fila 1 y un registro por fila. Solo existen en los archivos NEW.

Dos diferencias con el resto de extractores marcan todo el diseño de este módulo:

1. **No se usa `construir_grid`**: su tope de 250 filas dejaría fuera casi todo
   (`TD_datos_dia` llega a ~590 filas, `DATOS_MES` a ~1.275, y `BDP_datos_mes` supera las
   314.000). Las planas ni siquiera construyen grid: se recorren en streaming.

2. **`(en blanco)` se preserva como categoría real** en las etiquetas de fila de los
   pivots. Es un valor legítimo del negocio —una dimensión que no aplica—, no ruido. Por
   eso aquí se usa `_etiqueta_fila` en vez de `s()`, que lo convertiría en `None` y
   fusionaría filas distintas.

Las cabeceras se anclan **por contenido**, no por número de fila, porque el layout
difiere entre archivos NEW y STD.
"""

from __future__ import annotations

import datetime as dt
from collections.abc import Iterator
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from src.features.ingesta.celdas import NOISE, num, s, to_date
from src.features.ingesta.extractores.comunes import Grid, ResultadoExtractor

# Medidas que `BDP_datos_dia` despliega en filas (unpivot).
_MEDIDAS_BDP_DIA = ["VOLUMEN", "PORCENTAJE", "VOLDISMEZ", "VOL_ESTIMADO", "PROMEDIO"]

# Medidas de `BDP_datos_mes`: NO se despliegan (serían 3,1 M de filas); solo se guarda
# BPDEQ_M como valor y el resto se excluye de las dims.
_MEDIDAS_BDP_MES = [
    "VOLUMEN", "PORCENTAJE", "VOLDISMEZ", "BPD_M", "BPDA_AC", "BPDAC_5",
    "BPD_A", "BPDEQ_M", "BLSEQ", "BPDEQ_A",
]  # fmt: skip

# Columnas temporales que nunca son dimensión (la fecha va en su propio campo).
_COLUMNAS_TEMPORALES = {"FECHA", "MES", "AÑO", "ANO"}


def _etiqueta_fila(valor: Any) -> str | None:
    """Etiqueta de un pivot: preserva '(en blanco)'; solo vacío o `None` es ausencia.

    Deliberadamente distinto de `s()`, que trataría '(en blanco)' como ruido y lo
    convertiría en `None`, fusionando en una sola fila categorías que el negocio
    distingue.
    """
    if valor is None:
        return None
    texto = str(valor).strip()
    return None if texto == "" else texto


def _etiqueta_con_fecha_iso(valor: Any) -> str | None:
    """Como `_etiqueta_fila`, pero pasando las fechas descriptivas a ISO.

    Algunas columnas descriptivas de las hojas RAW traen fechas; sin normalizarlas, dos
    representaciones del mismo día generarían dims distintas y duplicarían la fila.
    """
    if valor is None:
        return None
    if isinstance(valor, (dt.date, dt.datetime)):
        return valor.isoformat()
    texto = str(valor).strip()
    return None if texto == "" else texto


def _etiqueta_sin_ruido(valor: Any) -> str | None:
    """Variante que además descarta el ruido de Excel — la usa `BDP_Programa`."""
    if valor is None:
        return None
    if isinstance(valor, (dt.date, dt.datetime)):
        return valor.isoformat()
    texto = str(valor).strip()
    return None if texto in NOISE else texto


def _grid_completo(hoja: Worksheet) -> tuple[Grid, int, int]:
    """Grid SIN tope de filas, con la última fila y columna ocupadas."""
    grid: Grid = {}
    ultima_fila = 0
    ultima_columna = 0
    for fila, valores in enumerate(hoja.iter_rows(values_only=True), start=1):
        for columna, valor in enumerate(valores, start=1):
            if valor is not None and str(valor).strip() != "":
                grid[(fila, columna)] = valor
                ultima_fila = max(ultima_fila, fila)
                ultima_columna = max(ultima_columna, columna)
    return grid, ultima_fila, ultima_columna


def _es_subtotal(valores: list[str | None]) -> bool:
    """True si alguna etiqueta de fila empieza por 'Total' — fila derivada, se excluye."""
    return any(v is not None and v.lower().startswith("total") for v in valores)


def _rellenar_niveles(
    actuales: list[str | None], arrastre: list[str | None]
) -> list[str | None]:
    """Propaga los niveles padre: un pivot solo escribe la etiqueta en la primera fila
    de cada grupo, y las siguientes la heredan."""
    for posicion, valor in enumerate(actuales):
        if valor is not None:
            arrastre[posicion] = valor
    return arrastre


# ── TD_datos_dia (tabla dinámica diaria) ─────────────────────────────────────

_NIVELES_TD = ["tipoproducto", "vice", "activos", "grupo1", "fuente"]
_GRUPOS_PRODUCCION = ("ECOPETROL", "SOCIOS")


def extraer_td_datos_dia(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'TD_datos_dia' → 1 tabla larga diaria, grano detalle.

    Cabeceras en tres filas superpuestas, ancladas por contenido:
    fila 19 = FECHA (se propaga desde la última fecha real), fila 20 = medida
    ('Suma de VOLDISMEZ', 'Suma de VOL_ESTIMADO'…), fila 21 = GRUPOPROD
    (ECOPETROL/SOCIOS), que es lo que define una columna de detalle.

    Se excluyen los subtotales: tanto las filas 'Total …' como las columnas de total por
    fecha y el gran total, que no llegan a `detalle` porque no tienen GRUPOPROD.
    """
    resultado = ResultadoExtractor(
        tablas_declaradas=[(1, "TD_datos_dia (detalle diario)")]
    )
    grid, ultima_fila, ultima_columna = _grid_completo(hoja)

    # Columnas de detalle: (columna, fecha, medida, grupoprod).
    detalle: list[tuple[int, dt.date, str, str]] = []
    fecha_vigente: dt.date | None = None
    medida_vigente: str | None = None
    for columna in range(6, ultima_columna + 1):
        fecha = to_date(grid.get((19, columna)))
        if fecha:
            fecha_vigente = fecha
        medida = s(grid.get((20, columna)))
        if medida and medida.lower().startswith("suma de "):
            medida_vigente = medida[8:].strip()
        grupo = s(grid.get((21, columna)))
        if grupo in _GRUPOS_PRODUCCION and fecha_vigente and medida_vigente:
            detalle.append((columna, fecha_vigente, medida_vigente, grupo))
    if not detalle:
        return resultado

    arrastre: list[str | None] = [None] * 5
    for fila in range(22, ultima_fila + 1):
        etiquetas = [_etiqueta_fila(grid.get((fila, c))) for c in range(1, 6)]
        if _es_subtotal(etiquetas):
            continue
        arrastre = _rellenar_niveles(etiquetas, arrastre)
        base = {
            _NIVELES_TD[i]: arrastre[i] for i in range(5) if arrastre[i] is not None
        }
        for columna, fecha, medida, grupo in detalle:
            valor = num(grid.get((fila, columna)))
            if valor is None:
                continue
            dims = dict(base)
            dims["grupoprod"] = grupo
            dims["medida"] = medida
            resultado.agregar(1, "TD_datos_dia (detalle diario)", dims, fecha, valor)
    return resultado


# ── DATOS_MES (tabla dinámica mensual) ───────────────────────────────────────

_NIVELES_DATOS_MES = ["escenario", "producto", "vice", "activos", "area", "campo"]


def extraer_datos_mes(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'DATOS_MES' → 1 tabla larga mensual, grano detalle.

    La fila de cabecera se localiza por contenido: es la que dice 'ESCENARIO' en la
    columna A. Las columnas A–F son los seis niveles de fila; desde la G van las fechas.

    La medida ('BPDEQ_M') y GRUPOPROD ('ECOPETROL') son filtros del pivot, no dimensiones:
    modelarlas como dims añadiría dos claves constantes a cada fila.
    """
    resultado = ResultadoExtractor(
        tablas_declaradas=[(1, "DATOS_MES (detalle mensual)")]
    )
    grid, ultima_fila, ultima_columna = _grid_completo(hoja)

    cabecera = None
    for fila in range(1, ultima_fila + 1):
        if (s(grid.get((fila, 1))) or "").upper() == "ESCENARIO":
            cabecera = fila
            break
    if cabecera is None:
        return resultado

    columnas_fecha = [
        (columna, fecha)
        for columna in range(7, ultima_columna + 1)
        if (fecha := to_date(grid.get((cabecera, columna)))) is not None
    ]
    if not columnas_fecha:
        return resultado

    arrastre: list[str | None] = [None] * 6
    for fila in range(cabecera + 1, ultima_fila + 1):
        etiquetas = [_etiqueta_fila(grid.get((fila, c))) for c in range(1, 7)]
        if _es_subtotal(etiquetas):
            continue
        arrastre = _rellenar_niveles(etiquetas, arrastre)
        base = {
            _NIVELES_DATOS_MES[i]: arrastre[i]
            for i in range(6)
            if arrastre[i] is not None
        }
        for columna, fecha in columnas_fecha:
            valor = num(grid.get((fila, columna)))
            if valor is None:
                continue
            resultado.agregar(
                1, "DATOS_MES (detalle mensual)", dict(base), fecha, valor
            )
    return resultado


# ── Hojas planas BDP_* ───────────────────────────────────────────────────────


def _cabecera_plana(
    filas: Iterator[tuple[Any, ...]],
) -> tuple[dict[int, str], dict[str, int]] | None:
    """Lee la fila 1 y devuelve (posición → nombre, NOMBRE → posición).

    El mapa por nombre es lo que hace estos extractores robustos a que Excel reordene
    columnas entre versiones del archivo: nada se lee por posición fija.
    """
    cabecera = next(filas, None)
    if cabecera is None:
        return None
    por_posicion = {
        posicion: nombre
        for posicion, valor in enumerate(cabecera)
        if (nombre := s(valor)) is not None
    }
    por_nombre = {nombre.upper(): posicion for posicion, nombre in por_posicion.items()}
    return por_posicion, por_nombre


def extraer_bdp_datos_dia(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'BDP_datos_dia' (solo NEW) → 1 tabla larga diaria por **unpivot** de 5 medidas.

    Emite una fila por cada (registro × medida con valor no nulo), conservando los ceros
    reales. Cada fila del origen es atómica: no hay subtotales ni etiquetas heredadas, así
    que aquí sí se usa `s()`.
    """
    resultado = ResultadoExtractor(
        tablas_declaradas=[(1, "BDP_datos_dia (detalle diario RAW)")]
    )
    filas = hoja.iter_rows(values_only=True)
    cabeceras = _cabecera_plana(filas)
    if cabeceras is None:
        return resultado
    por_posicion, por_nombre = cabeceras

    if "FECHA" not in por_nombre or not any(m in por_nombre for m in _MEDIDAS_BDP_DIA):
        return resultado

    posicion_fecha = por_nombre["FECHA"]
    excluidas = _COLUMNAS_TEMPORALES | set(_MEDIDAS_BDP_DIA)
    columnas_dim = [
        (posicion, por_posicion[posicion].lower())
        for posicion in sorted(por_posicion)
        if por_posicion[posicion].upper() not in excluidas
    ]
    columnas_medida = [(por_nombre[m], m) for m in _MEDIDAS_BDP_DIA if m in por_nombre]

    for fila in filas:
        fecha = to_date(fila[posicion_fecha]) if posicion_fecha < len(fila) else None
        if fecha is None:
            continue
        base: dict[str, Any] = {}
        for posicion, clave in columnas_dim:
            if posicion < len(fila) and (valor_dim := s(fila[posicion])) is not None:
                base[clave] = valor_dim
        for posicion, nombre_medida in columnas_medida:
            valor = num(fila[posicion]) if posicion < len(fila) else None
            if valor is None:
                continue
            dims = dict(base)
            dims["medida"] = nombre_medida
            resultado.agregar(
                1, "BDP_datos_dia (detalle diario RAW)", dims, fecha, valor
            )
    return resultado


def extraer_bdp_datos_mes(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'BDP_datos_mes' (solo NEW) → 1 tabla larga mensual, **una fila por registro**.

    A diferencia de `BDP_datos_dia`, aquí **no** se despliegan las medidas: son 10 y la
    hoja tiene ~314.000 registros, así que el unpivot daría más de 3 millones de filas.
    Se guarda `BPDEQ_M` como valor —la misma medida del pivot DATOS_MES— y las otras nueve
    quedan fuera.

    Es un cubo multi-año (97 meses, 4 escenarios), de ahí su tamaño.
    """
    resultado = ResultadoExtractor(
        tablas_declaradas=[(1, "BDP_datos_mes (detalle mensual RAW)")]
    )
    filas = hoja.iter_rows(values_only=True)
    cabeceras = _cabecera_plana(filas)
    if cabeceras is None:
        return resultado
    por_posicion, por_nombre = cabeceras

    if "FECHA" not in por_nombre or "BPDEQ_M" not in por_nombre:
        return resultado

    posicion_fecha = por_nombre["FECHA"]
    posicion_valor = por_nombre["BPDEQ_M"]
    excluidas = _COLUMNAS_TEMPORALES | set(_MEDIDAS_BDP_MES)
    columnas_dim = [
        (posicion, por_posicion[posicion].lower())
        for posicion in sorted(por_posicion)
        if por_posicion[posicion].upper() not in excluidas
    ]

    for fila in filas:
        fecha = to_date(fila[posicion_fecha]) if posicion_fecha < len(fila) else None
        if fecha is None:
            continue
        valor = num(fila[posicion_valor]) if posicion_valor < len(fila) else None
        dims: dict[str, Any] = {}
        for posicion, clave in columnas_dim:
            if posicion < len(fila):
                etiqueta = _etiqueta_con_fecha_iso(fila[posicion])
                if etiqueta is not None:
                    dims[clave] = etiqueta
        resultado.agregar(1, "BDP_datos_mes (detalle mensual RAW)", dims, fecha, valor)
    return resultado


def extraer_bdp_programa(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'BDP_Programa' (solo NEW) → 1 tabla larga, una fila por registro.

    `valor` es Volumen (la cuota ECP programada). Las otras 12 columnas se conservan como
    dimensiones —incluidas `Produccion_total` y `Part_ECP`— por decisión del usuario: no
    se pierde ninguna de las 14 columnas del origen.

    Las claves de dimensión se derivan del encabezado en minúsculas con los espacios
    convertidos en guion bajo ('Fecha Version' → 'fecha_version').
    """
    resultado = ResultadoExtractor(
        tablas_declaradas=[(1, "BDP_Programa (programa RAW)")]
    )
    filas = hoja.iter_rows(values_only=True)
    cabeceras = _cabecera_plana(filas)
    if cabeceras is None:
        return resultado
    por_posicion, por_nombre = cabeceras

    if "FECHA" not in por_nombre or "VOLUMEN" not in por_nombre:
        return resultado

    posicion_fecha = por_nombre["FECHA"]
    posicion_valor = por_nombre["VOLUMEN"]
    columnas_dim = [
        (posicion, por_posicion[posicion].lower().replace(" ", "_"))
        for posicion in sorted(por_posicion)
        if por_posicion[posicion].upper() not in {"FECHA", "VOLUMEN"}
    ]

    for fila in filas:
        fecha = to_date(fila[posicion_fecha]) if posicion_fecha < len(fila) else None
        if fecha is None:
            continue
        valor = num(fila[posicion_valor]) if posicion_valor < len(fila) else None
        dims: dict[str, Any] = {}
        for posicion, clave in columnas_dim:
            if posicion < len(fila):
                etiqueta = _etiqueta_sin_ruido(fila[posicion])
                if etiqueta is not None:
                    dims[clave] = etiqueta
        resultado.agregar(1, "BDP_Programa (programa RAW)", dims, fecha, valor)
    return resultado
