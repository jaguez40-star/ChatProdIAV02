"""Extractores de las hojas de reporte — portados literalmente del sistema viejo.

Agrupa las hojas que publican cifras ya calculadas para consumo humano: la bitácora, el
programa, los reportes de WhatsApp/DPP/presidencia y el cálculo de trimestre.

Todas comparten un rasgo: **ingieren los valores cacheados tal como aparecen**, sin
recalcular. Son reportes derivados y, por requerimiento del usuario, su cifra publicada es
la que manda aunque se pudiera recomputar desde las hojas crudas.
"""

from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from src.features.ingesta.celdas import num, s, to_date
from src.features.ingesta.extractores.comunes import (
    Grid,
    ResultadoExtractor,
    construir_grid,
    meses_contiguos,
)

# ── (Bitacora) ───────────────────────────────────────────────────────────────

_PRODUCTOS_BITACORA: frozenset[str] = frozenset({"CRUDO", "GAS", "BLANCOS"})
_VICES_BITACORA: frozenset[str] = frozenset({"VRC", "VRO", "VAO", "VFS", "VPI", "VEX"})

_TABLAS_BITACORA = [
    (1, "Tabla 1 (REAL)"),
    (2, "Tabla 2 (PROGRAMA)"),
    (3, "Tabla 3 (PROYECCIÓN)"),
]


def _bloque_de(etiqueta: str) -> tuple[int, str] | None:
    """A qué tabla pertenece un bloque, por el texto de su título '*** … ***'."""
    mayus = etiqueta.upper()
    if "REAL" in mayus:
        return (1, "Tabla 1 (REAL)")
    if "PROGRAMA" in mayus:
        return (2, "Tabla 2 (PROGRAMA)")
    if "PROYEC" in mayus:
        return (3, "Tabla 3 (PROYECCIÓN)")
    return None


def extraer_bitacora(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja '(Bitacora)' → 3 tablas TIPOPRODUCTO × VICE por fecha.

    Declara siempre las 3 tablas para que el front muestre tres ítems aunque PROGRAMA
    venga con `#N/A` (ocurre en los archivos STD, que producen 0 filas ahí).

    Excluye los subtotales —se reconocen porque su columna VICE viene vacía: 'Total CRUDO',
    'Total general'…— y la columna agregada mensual a la derecha de las fechas, que se
    descarta sola porque `to_date` no la reconoce como fecha.

    El producto se rellena hacia abajo: la columna A solo lo escribe en la primera fila de
    cada grupo.
    """
    filas_hoja = [list(fila) for fila in hoja.iter_rows(values_only=True)]
    total_filas = len(filas_hoja)
    resultado = ResultadoExtractor(tablas_declaradas=list(_TABLAS_BITACORA))

    def texto(fila: int, columna: int) -> str:
        if 0 <= fila < total_filas and 0 <= columna < len(filas_hoja[fila]):
            return s(filas_hoja[fila][columna]) or ""
        return ""

    fila = 0
    while fila < total_filas:
        titulo = texto(fila, 0)
        if not titulo.startswith("***"):
            fila += 1
            continue
        bloque = _bloque_de(titulo)
        if bloque is None:
            fila += 1
            continue
        indice, etiqueta = bloque

        # El encabezado de fechas es la fila cuya columna A dice 'TIPOPRODUCTO'.
        cabecera = fila + 1
        while (
            cabecera < total_filas
            and texto(cabecera, 0).upper() != "TIPOPRODUCTO"
            and not texto(cabecera, 0).startswith("***")
        ):
            cabecera += 1
        if cabecera >= total_filas or texto(cabecera, 0).upper() != "TIPOPRODUCTO":
            fila += 1
            continue

        # Desde la columna C; la columna 'REAL' (agregado mensual) no es fecha → None.
        fechas = [to_date(v) for v in filas_hoja[cabecera][2:]]
        producto: str | None = None
        actual = cabecera + 1
        while actual < total_filas and not texto(actual, 0).startswith("***"):
            columna_a = texto(actual, 0).upper()
            if columna_a in _PRODUCTOS_BITACORA:
                producto = columna_a  # forward-fill: la columna A viene dispersa
            vice = texto(actual, 1).upper()
            # Solo es fila de datos si tiene VICE — así se excluyen los subtotales.
            if vice in _VICES_BITACORA and producto:
                for posicion, valor_celda in enumerate(filas_hoja[actual][2:]):
                    fecha = fechas[posicion] if posicion < len(fechas) else None
                    valor = num(valor_celda)
                    if fecha is None or valor is None:
                        continue  # fecha inválida, #N/A o celda vacía
                    resultado.agregar(
                        indice,
                        etiqueta,
                        {"tipoproducto": producto, "vice": vice},
                        fecha,
                        valor,
                    )
            actual += 1
        fila = actual

    return resultado


# ── PROGRAMA ─────────────────────────────────────────────────────────────────

# La hoja PROGRAMA es más alta que las demás: sus tablas llegan a la fila 325.
MAX_FILAS_PROGRAMA = 330

_TABLAS_PROGRAMA = [
    (1, "Tabla 1 (Produccion_total)"),
    (2, "Tabla 2 (Volumen)"),
    (3, "Tabla 3 (Programa por VICE)"),
    (4, "Tabla 4 (verificador carga)"),
]

# (tabla, etiqueta, [(columna, nombre_dim)], fila_encabezado, fila_inicial, fila_final)
_ESPECIFICACIONES_PROGRAMA: list[
    tuple[int, str, list[tuple[int, str]], int, int, int]
] = [
    (
        1,
        "Tabla 1 (Produccion_total)",
        [(1, "tipoproducto"), (2, "area"), (3, "campo")],
        5,
        6,
        312,
    ),
    (2, "Tabla 2 (Volumen)", [(87, "producto"), (88, "vice")], 5, 6, 24),
    (
        3,
        "Tabla 3 (Programa por VICE)",
        [(87, "tipoproducto"), (88, "vice")],
        29,
        31,
        52,
    ),
    (
        4,
        "Tabla 4 (verificador carga)",
        [(179, "producto"), (180, "vice"), (181, "area"), (182, "campo")],
        10,
        11,
        325,
    ),
]


def _columnas_fecha(
    grid: Grid, fila_encabezado: int, columna_inicial: int
) -> list[tuple[int, Any]]:
    """Columnas-fecha desde `columna_inicial`, saltando las que no lo son.

    A diferencia de `meses_contiguos`, aquí NO se corta en la primera no-fecha: la hoja
    intercala columnas '(en blanco)' y 'Total general' entre las de fecha. Se corta tras
    **tres columnas vacías seguidas**, que es lo que separa una tabla de la siguiente.
    """
    columnas: list[tuple[int, Any]] = []
    columna = columna_inicial
    vacias_seguidas = 0
    while vacias_seguidas < 3 and columna < columna_inicial + 280:
        valor = grid.get((fila_encabezado, columna))
        if valor is None:
            vacias_seguidas += 1
        else:
            vacias_seguidas = 0
            fecha = to_date(valor)
            if fecha is not None:
                columnas.append((columna, fecha))
        columna += 1
    return columnas


def extraer_programa(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'PROGRAMA' → 4 tablas de valores cacheados.

    **No recalcula desde `BDP_Programa`** (decisión del usuario, 2026-06-30): se ingiere
    lo que la hoja publica.

    Las filas de total SÍ se incluyen; lo que se omite son las columnas sin fecha.

    El relleno hacia abajo de las dimensiones es **consciente de los subtotales**: cuando
    una celda dice 'Total…', las columnas a su derecha quedan en blanco y NO heredan el
    valor de la fila anterior — heredarlo etiquetaría el subtotal con el último campo
    concreto, inventando un dato que la hoja no dice.
    """
    grid, _ = construir_grid(hoja, max_filas=MAX_FILAS_PROGRAMA)
    resultado = ResultadoExtractor(tablas_declaradas=list(_TABLAS_PROGRAMA))

    for (
        indice,
        etiqueta,
        columnas_dim,
        cabecera,
        desde,
        hasta,
    ) in _ESPECIFICACIONES_PROGRAMA:
        columnas_fecha = _columnas_fecha(grid, cabecera, columnas_dim[-1][0] + 1)
        if not columnas_fecha:
            continue
        arrastre: dict[int, str] = {}
        for fila in range(desde, hasta + 1):
            crudas = [grid.get((fila, columna)) for columna, _ in columnas_dim]
            if all(v is None or str(v).strip() == "" for v in crudas):
                continue

            valores: list[str | None] = []
            hubo_total = False
            for posicion, valor_crudo in enumerate(crudas):
                texto_dim = s(valor_crudo)
                if texto_dim:
                    if texto_dim.lower().startswith("total"):
                        hubo_total = True
                    arrastre[posicion] = texto_dim
                    valores.append(texto_dim)
                else:
                    valores.append(None if hubo_total else arrastre.get(posicion))

            dims = {
                nombre: valores[posicion]
                for posicion, (_, nombre) in enumerate(columnas_dim)
                if valores[posicion] is not None
            }
            for columna, fecha in columnas_fecha:
                valor = num(grid.get((fila, columna)))
                if valor is None:
                    continue
                resultado.agregar(indice, etiqueta, dims, fecha, valor)

    return resultado


# ── Reporte Whatsapp ─────────────────────────────────────────────────────────

_TABLAS_WHATSAPP = [
    (1, "T1 PROGRAMA (consolidado)"),
    (2, "T2 Mes en curso (consolidado)"),
    (3, "T3 Proyección mes (consolidado)"),
    (4, "T4 Trimestre (consolidado)"),
    (5, "T5 YTD (consolidado)"),
    (6, "T6 Año (consolidado)"),
    (7, "T7 Crudo por activo (izq L-P)"),
    (8, "T8 Gas por activo (izq L-P)"),
    (9, "T9 Equivalente por filial (izq L-P)"),
    (10, "T10 Crudo por activo (der S-T)"),
    (11, "T11 Gas por activo (der S-T)"),
    (12, "T12 Equivalente por filial (der S-T)"),
]

# Rótulos que identifican la fila-título de un bloque consolidado.
_PALABRAS_ENCABEZADO = ("real", "plan", "proy", "delta", "programa", "pop")
_ORDEN_SEGMENTOS = ["Ecopetrol", "Filiales", "Upstream"]

_METRICAS_IZQUIERDA = [(13, "M"), (14, "N"), (15, "O"), (16, "P")]
_METRICAS_DERECHA = [(19, "S"), (20, "T")]
# Orden estable de las secciones por activo: 1=Crudo, 2=Gas, 3=Equivalente.
_SECCIONES_ACTIVO = [(7, 10), (8, 11), (9, 12)]


def extraer_whatsapp(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'Reporte Whatsapp' → 12 tablas de valores cacheados.

    **Bloque C–F**: 6 tablas consolidadas apiladas (Ecopetrol/Filiales/Upstream ×
    Crudo/Gas/Blancos), con 3 métricas por fila (columnas D/E/F).

    ⚠️ Los títulos de la columna C **varían entre archivos** (Junio, Febrero, 1Q, 4Q…),
    así que las etiquetas de tabla son **posicionales y estables**: el índice 1..6 se
    asigna por orden de aparición, no por el texto. Si se usara el texto, cada mes
    produciría tablas con nombres distintos.

    **Bloque L–T**: 6 tablas por activo. Las columnas Q/R (vacías) separan la mitad
    izquierda (4 métricas M/N/O/P) de la derecha (2 métricas S/T). Las secciones se anclan
    por la celda 'ACTIVOS' en la columna L; los blancos internos **no** cortan la sección
    —hacen falta tres seguidos—, y cierra en la nota '**…' o en el siguiente 'ACTIVOS'.
    La sección Equivalente puede faltar en archivos STD: sus tablas se declaran igual.

    `fecha=None` a propósito: la celda 'Producción al:' no es fiable (en archivos STD
    antiguos trae valores recalculados de otra época). El linaje temporal lo aporta
    `reporte_id` → `config_reporte`.

    Las filas de total y subtotal **sí** se incluyen, por decisión del dueño del reporte.
    """
    grid, ultima_fila = construir_grid(hoja)
    resultado = ResultadoExtractor(tablas_declaradas=list(_TABLAS_WHATSAPP))
    etiquetas = dict(_TABLAS_WHATSAPP)

    # ── Bloque C–F: 6 tablas consolidadas apiladas ──────────────────────────
    titulos = [
        fila
        for fila in range(1, ultima_fila + 1)
        if s(grid.get((fila, 3)))
        and (encabezado := s(grid.get((fila, 4))))
        and any(clave in encabezado.lower() for clave in _PALABRAS_ENCABEZADO)
    ]
    for posicion, fila_titulo in enumerate(titulos[:6]):
        indice = posicion + 1  # índice posicional ESTABLE (1..6)
        metricas = [
            (columna, letra, nombre)
            for columna, letra in ((4, "D"), (5, "E"), (6, "F"))
            if (nombre := s(grid.get((fila_titulo, columna))))
        ]
        segmento_actual = 0
        fila = fila_titulo + 1
        while fila <= ultima_fila:
            etiqueta_fila = s(grid.get((fila, 3)))
            if etiqueta_fila is None:
                break  # un blanco en la columna C cierra la tabla
            mayus = etiqueta_fila.upper()
            if mayus == "ECOPETROL":
                segmento, concepto, segmento_actual = "Ecopetrol", "Total", 1
            elif mayus == "FILIALES":
                segmento, concepto, segmento_actual = "Filiales", "Total", 2
            elif mayus == "UPSTREAM":
                segmento, concepto = "Upstream", "Total"
            else:
                segmento, concepto = _ORDEN_SEGMENTOS[segmento_actual], etiqueta_fila

            for columna, letra, nombre in metricas:
                valor = num(grid.get((fila, columna)))
                if valor is None:
                    continue
                resultado.agregar(
                    indice,
                    etiquetas[indice],
                    {
                        "segmento": segmento,
                        "concepto": concepto,
                        "columna": letra,
                        "metrica": nombre,
                    },
                    None,
                    valor,
                )
            if mayus == "UPSTREAM":
                break  # Upstream cierra la tabla
            fila += 1

    # ── Bloque L–T: secciones ancladas por 'ACTIVOS' en la columna L ────────
    encabezados = [
        fila
        for fila in range(1, ultima_fila + 1)
        if s(grid.get((fila, 12))) == "ACTIVOS"
    ]
    for posicion, fila_encabezado in enumerate(encabezados[:3]):
        indice_izq, indice_der = _SECCIONES_ACTIVO[posicion]
        # El nombre de la métrica puede faltar; se anota explícito para que ambas listas
        # tengan el mismo tipo al iterarlas juntas más abajo.
        metricas_izq: list[tuple[int, str, str | None]] = [
            (columna, letra, s(grid.get((fila_encabezado, columna))))
            for columna, letra in _METRICAS_IZQUIERDA
        ]
        metricas_der: list[tuple[int, str, str | None]] = [
            (columna, letra, s(grid.get((fila_encabezado, columna))))
            for columna, letra in _METRICAS_DERECHA
        ]
        fila = fila_encabezado + 1
        blancos = 0
        while fila <= ultima_fila:
            activo = s(grid.get((fila, 12)))
            if activo is None:
                blancos += 1
                if blancos >= 3:  # tres blancos seguidos = fin de sección
                    break
                fila += 1
                continue
            if activo.startswith("**") or activo == "ACTIVOS":
                break  # nota al pie o siguiente encabezado
            blancos = 0
            for indice, metricas_seccion in (
                (indice_izq, metricas_izq),
                (indice_der, metricas_der),
            ):
                for columna, letra, nombre in metricas_seccion:
                    valor = num(grid.get((fila, columna)))
                    if valor is None:
                        continue
                    resultado.agregar(
                        indice,
                        etiquetas[indice],
                        {"activo": activo, "columna": letra, "metrica": nombre},
                        None,
                        valor,
                    )
            fila += 1

    return resultado


# ── Reporte DPP ──────────────────────────────────────────────────────────────

_TABLAS_DPP = [
    (1, "COMPARATIVO DÍA"),
    (2, "PROYECCIÓN MES"),
    (3, "CUMPLIMIENTO MES"),
    (4, "PROYECCIÓN AÑO"),
    (5, "CUMPLIMIENTO AÑO"),
]

# (fila de Excel, etiqueta con el grupo embebido). Filas fijas, verificadas en 3 archivos.
# Las filas de total se nombran 'TOTAL …' para que el visor las resalte.
_FILAS_DPP = [
    (13, "ECOPETROL · CRUDO"),
    (14, "ECOPETROL · GAS"),
    (15, "ECOPETROL · BLANCOS"),
    (16, "ECOPETROL · ECP EXPLORACIÓN"),
    (17, "TOTAL ECOPETROL"),
    (18, "FILIALES · CRUDO"),
    (19, "FILIALES · GAS"),
    (20, "FILIALES · BLANCOS"),
    (21, "TOTAL FILIALES"),
    (22, "TOTAL UPSTREAM"),
]

_COLUMNAS_DPP: list[tuple[int, list[tuple[int, str]]]] = [
    (
        1,
        [
            (3, "REAL día anterior"),
            (4, "PROGRAMA día anterior"),
            (5, "REAL día actual"),
            (6, "PROGRAMA día actual"),
        ],
    ),
    (2, [(8, "REAL mes"), (9, "PROYECCIÓN MES")]),
    (
        3,
        [
            (11, "P50 (META) mes"),
            (12, "DIFERENCIA mes"),
            (13, "PRODUCCIÓN NECESARIA mes"),
        ],
    ),
    (4, [(15, "REAL año (YTD)"), (16, "PROYECCIÓN AÑO")]),
    (
        5,
        [
            (18, "P50 (META) año"),
            (19, "DIFERENCIA año"),
            (20, "PRODUCCIÓN NECESARIA año"),
        ],
    ),
]


def extraer_dpp(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'Reporte DPP' → 5 tablas matriciales (segmento × métrica), `fecha=None`.

    Snapshot de KPIs comparativos; **no recalcula**.

    El grupo se embebe en `dims.fila` ('ECOPETROL · CRUDO') porque los segmentos se
    repiten entre bloques: sin el prefijo, CRUDO de Ecopetrol y CRUDO de Filiales
    colisionarían en la misma clave.

    **Regla de celdas propia de DPP** (acordada 2026-06-30, distinta de la regla general de
    ruido): una celda con un **error de Excel** (`#¡REF!`, `#N/A`…) se inserta con
    `valor=None` —preserva la forma de la tabla, deja el hueco visible—, mientras que una
    celda **vacía de verdad** se salta. Por eso aquí no se filtra por `num() is None`.
    """
    resultado = ResultadoExtractor(tablas_declaradas=list(_TABLAS_DPP))
    etiquetas = dict(_TABLAS_DPP)

    # Se leen las celdas crudas (sin `construir_grid`): hace falta distinguir la celda
    # vacía del error de Excel, y `construir_grid` descarta ambas por igual.
    grid: dict[tuple[int, int], Any] = {}
    for fila, valores in enumerate(hoja.iter_rows(values_only=True), start=1):
        if fila > 22:
            break
        for columna, valor in enumerate(valores, start=1):
            grid[(fila, columna)] = valor

    for indice, columnas in _COLUMNAS_DPP:
        for fila_excel, etiqueta_fila in _FILAS_DPP:
            for columna, etiqueta_columna in columnas:
                crudo = grid.get((fila_excel, columna))
                if crudo is None or str(crudo).strip() == "":
                    continue  # celda vacía de verdad → se salta
                # Error de Excel (no numérico, no vacío) → se inserta como None.
                resultado.agregar(
                    indice,
                    etiquetas[indice],
                    {"fila": etiqueta_fila, "columna": etiqueta_columna},
                    None,
                    num(crudo),
                )

    return resultado


# ── REPORTE_PRESIDENT ────────────────────────────────────────────────────────

_TABLAS_PRESIDENT = [
    (1, "Tabla 1 (Producción día)"),
    (2, "Tabla 2 (Producción mes)"),
]

# Medidas por POSICIÓN relativa al ancla: el sufijo de mes ('Mar', 'Dic') y el rótulo
# 'P50'/'Reto NNNK' del encabezado VARÍAN entre archivos, así que el texto no sirve de clave.
_MEDIDAS_MES = [
    (-2, "real_mes"),
    (-1, "proy_mes"),
    (0, "base_p50"),
    (1, "delta_p50"),
    (2, "compromiso"),
    (3, "delta_compromiso"),
]
_MEDIDAS_DIA = [(0, "real_dia"), (1, "programa_dia"), (2, "delta_dia")]


def extraer_reporte_president(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'REPORTE_PRESIDENT' → 2 tablas (día y mes), grano entidad × medida.

    Se ancla en la celda **'Base P50'**, no en una fila fija. Dos bloques comparten la
    fila de encabezado, cada uno con su propia columna de etiqueta de entidad: el bloque
    DÍA (B:E) y el bloque MES (G:M).

    La escala es kbpe (mundo corporativo P50), **no** la del fact diario.

    Aporta dos cosas que ninguna otra hoja da: BLANCOS con Real/Proy/P50, y el
    «compromiso» (la columna L trae el Reto cuando difiere del P50, y el P50 cuando no hay
    stretch ese periodo).

    El bloque DÍA suele venir en `#REF!`/`#N/A` en varios archivos; `num()` lo descarta y
    esa medida simplemente no emite fila. El bloque MES es el fiable.

    Queda fuera el texto narrativo ('Principales eventos', 'Actividades programadas'):
    no es numérico.
    """
    grid: Grid = {}
    for fila, valores in enumerate(
        hoja.iter_rows(min_row=1, max_row=50, max_col=14, values_only=True), start=1
    ):
        for columna, valor in enumerate(valores, start=1):
            if valor is not None and str(valor).strip() != "":
                grid[(fila, columna)] = valor

    resultado = ResultadoExtractor(tablas_declaradas=list(_TABLAS_PRESIDENT))

    def buscar_celda(texto_buscado: tuple[str, ...]) -> tuple[int, int] | None:
        for (fila, columna), valor in grid.items():
            if isinstance(valor, str) and valor.strip().lower() in texto_buscado:
                return fila, columna
        return None

    ancla = buscar_celda(("base p50",))
    if ancla is None:
        return resultado
    fila_encabezado, columna_base = ancla

    dia = buscar_celda(("real día", "real dia"))
    columna_dia = dia[1] if (dia and dia[0] == fila_encabezado) else None

    # La etiqueta de entidad del bloque MES está tres columnas a la izquierda del ancla.
    columna_entidad = columna_base - 3
    entidades: list[tuple[int, str]] = []
    fila = fila_encabezado + 1
    while (entidad := s(grid.get((fila, columna_entidad)))) is not None:
        entidades.append((fila, entidad))
        fila += 1

    for fila, entidad in entidades:
        for desplazamiento, medida in _MEDIDAS_MES:
            valor = num(grid.get((fila, columna_base + desplazamiento)))
            if valor is None:
                continue
            resultado.agregar(
                2,
                "Tabla 2 (Producción mes)",
                {"entidad": entidad, "medida": medida},
                None,
                valor,
            )
        if columna_dia:
            entidad_dia = s(grid.get((fila, columna_dia - 1))) or entidad
            for desplazamiento, medida in _MEDIDAS_DIA:
                valor = num(grid.get((fila, columna_dia + desplazamiento)))
                if valor is None:
                    continue
                resultado.agregar(
                    1,
                    "Tabla 1 (Producción día)",
                    {"entidad": entidad_dia, "medida": medida},
                    None,
                    valor,
                )

    return resultado


# ── CÁLCULO DE TRIMESTRE ─────────────────────────────────────────────────────

_TABLAS_TRIMESTRE = [
    (1, "PROGRAMA MES + FILIALES"),
    (2, "PROYECCIÓN AÑO producto×mes"),
    (3, "PROYECCIÓN AÑO producto×empresa×mes"),
    (4, "GRUPO EMPRESA×mes"),
    (5, "P50 631 PLAN (trimestres)"),
    (6, "P50 631 REAL (trimestres)"),
    (7, "P50 621,9 PLAN (trimestres)"),
    (8, "POP PLAN (trimestres)"),
]

_ETIQUETAS_TRIMESTRE = ["1Q", "2Q", "3Q", "4Q"]

# (tabla, columna_etiqueta, columnas de trimestre, rango de filas, embebe el bloque)
_ESPECIFICACIONES_TRIMESTRE: list[tuple[int, int, list[int], range, bool]] = [
    (5, 1, [2, 3, 4, 5], range(55, 67), True),
    (6, 8, [9, 10, 11, 12], range(55, 69), True),
    (7, 1, [2, 3, 4, 5], range(77, 83), False),
    (8, 1, [2, 3, 4, 5], range(88, 95), False),
]

# Frontera entre el bloque ECP y el de FILIALES en las tablas 5 y 6.
_FILA_FRONTERA_FILIALES = 62


def extraer_calculo_trimestre(hoja: Worksheet) -> ResultadoExtractor:
    """Hoja 'CÁLCULO DE TRIMESTRE' → 8 tablas. Hoja intermedia, heterogénea; valores tal cual.

    - **T1** (D-I, filas 8-22): matriz snapshot con dos bloques — PROGRAMA MES
      (producto × valor) y FILIALES (producto × empresa). `fecha=None`.
    - **T2/T3/T4**: temporales por mes; los meses salen del encabezado contiguo desde la
      columna C, que corta antes de 'Promedio Año'.
    - **T5-T8**: trimestrales (columna = 1Q..4Q, fila = concepto).

    En T5/T6 el mismo concepto aparece dos veces —una en el bloque ECP y otra en el de
    FILIALES—, así que a las etiquetas **que se repiten** se les añade el bloque:
    'CRUDO (ECP)' y 'CRUDO (FILIALES)'. Solo a las repetidas: sufijar todas ensuciaría
    etiquetas que ya son únicas (VDP, TOTAL, UPSTREAM…).
    """
    grid, _ = construir_grid(hoja)
    resultado = ResultadoExtractor(tablas_declaradas=list(_TABLAS_TRIMESTRE))
    etiquetas = dict(_TABLAS_TRIMESTRE)

    def emitir(indice: int, dims: dict[str, Any], fecha: Any, crudo: Any) -> None:
        valor = num(crudo)
        if valor is None:
            return
        resultado.agregar(indice, etiquetas[indice], dims, fecha, valor)

    # ── T1: matriz snapshot (columnas D-I, filas 8-22) ──────────────────────
    for fila in range(8, 23):
        programa = s(grid.get((fila, 4)))  # D: producto
        if programa is not None:
            emitir(
                1,
                {"fila": programa, "columna": "Programa mes"},
                None,
                grid.get((fila, 5)),
            )
        filial = s(grid.get((fila, 7)))  # G: producto
        if filial is not None:
            empresa = s(grid.get((fila, 8)))  # H: empresa
            etiqueta_fila = f"{filial} · {empresa}" if empresa else filial
            emitir(
                1,
                {"fila": etiqueta_fila, "columna": "Filiales"},
                None,
                grid.get((fila, 9)),
            )

    # ── T2/T3/T4: temporales (meses desde la columna C) ─────────────────────
    for fila in range(25, 32):  # T2: producto × mes
        producto = s(grid.get((fila, 1)))
        if producto is None:
            continue
        for columna, fecha in meses_contiguos(grid, 24, 3):
            emitir(2, {"producto": producto}, fecha, grid.get((fila, columna)))

    for fila in range(36, 49):  # T3: producto × empresa × mes
        producto = s(grid.get((fila, 1)))
        if producto is None:
            continue
        dims: dict[str, Any] = {"producto": producto}
        empresa = s(grid.get((fila, 2)))
        if empresa is not None:
            dims["empresa"] = empresa
        for columna, fecha in meses_contiguos(grid, 35, 3):
            emitir(3, dims, fecha, grid.get((fila, columna)))

    for fila in range(51, 52):  # T4: GRUPO EMPRESA × mes
        concepto = s(grid.get((fila, 1)))
        if concepto is None:
            continue
        for columna, fecha in meses_contiguos(grid, 50, 3):
            emitir(4, {"concepto": concepto}, fecha, grid.get((fila, columna)))

    # ── T5-T8: trimestrales ─────────────────────────────────────────────────
    for (
        indice,
        columna_etiqueta,
        columnas_q,
        rango,
        embebe_bloque,
    ) in _ESPECIFICACIONES_TRIMESTRE:
        etiquetas_filas = [s(grid.get((fila, columna_etiqueta))) for fila in rango]
        repetidas = (
            {
                etiqueta
                for etiqueta in etiquetas_filas
                if etiqueta is not None and etiquetas_filas.count(etiqueta) > 1
            }
            if embebe_bloque
            else set()
        )
        for fila in rango:
            concepto_fila = s(grid.get((fila, columna_etiqueta)))
            if concepto_fila is None:
                continue
            if concepto_fila in repetidas:
                bloque = "ECP" if fila < _FILA_FRONTERA_FILIALES else "FILIALES"
                concepto_fila = f"{concepto_fila} ({bloque})"
            for columna, trimestre in zip(
                columnas_q, _ETIQUETAS_TRIMESTRE, strict=False
            ):
                emitir(
                    indice,
                    {"fila": concepto_fila, "columna": trimestre},
                    None,
                    grid.get((fila, columna)),
                )

    return resultado
