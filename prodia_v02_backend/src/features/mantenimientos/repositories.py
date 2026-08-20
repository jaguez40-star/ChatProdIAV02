"""Lectura de `Eventos_OW.xlsx` — eventos de servicio a pozo.

Portado de `routes/api.py:414-465` del sistema viejo.

**A1 — singleton bajo lock con doble chequeo.** El parseo tarda ~1,53 s y el
servidor atiende con hilos: sin el lock, N peticiones concurrentes (el prefetch
del login dispara varias) parsearían el archivo N veces.

**A2 — `FinalizaEvento` vacío significa evento ABIERTO, no fila inválida.** Son
3.305 de 6.850 filas (48 %). Descartarlas perdía justo los eventos que siguen
corriendo, que son los que importan.

Cero I/O en tiempo de import (AP-2): el archivo se abre en la primera llamada.
"""

from __future__ import annotations

import threading
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import TypedDict

from src.core.config import get_settings
from src.core.logger import get_logger

logger = get_logger("mantenimientos.repositories")


class EventoOW(TypedDict):
    """Una fila del archivo, ya normalizada."""

    campo: str
    pozo: str
    tipo: str
    inicio: datetime
    fin: datetime | None  # `None` = evento ABIERTO (A2)


# Columnas del .xlsx (Sheet1), por posición — el archivo no tiene cabeceras
# estables. `tipo` cae en cascada: Objetivo Ppal → TIPO_EVENTO → Evento.
_COL_CAMPO = 3
_COL_POZO = 4
_COL_INICIO = 7
_COL_FIN = 8
_COLS_TIPO = (12, 10, 6)

# 5 filas traen el año mal tecleado (2526/2626/3026/2016). Se tratan como
# abiertas en vez de descartarse: el evento existió, solo su cierre es basura.
_ANIO_MIN, _ANIO_MAX = 2020, 2030

_cache_eventos: list[EventoOW] | None = None
_lock_carga = threading.Lock()


def normalizar(texto: str | None) -> str:
    """UPPER + trim + pliega acentos y ñ.

    MISMO criterio que `catalogo_entidades.norm`. El `.xlsx` viene en NFC
    ('CAÑO SUR ESTE') y nadie normaliza del lado de Postgres: un match literal
    sería byte a byte y una fuente en NFD rompería campos EN SILENCIO.
    Verificado que hoy da el mismo resultado que el literal (92 campos = 92),
    así que es blindaje sin cambio de conducta.
    """
    descompuesto = unicodedata.normalize("NFKD", (texto or "").strip().upper())
    sin_tildes = "".join(c for c in descompuesto if not unicodedata.combining(c))
    return " ".join(sin_tildes.split())


def ruta_eventos() -> Path | None:
    """Ruta configurada, o `None` si falta el archivo.

    `None` es un estado normal: el archivo vive fuera del repo y puede no estar
    en una máquina recién clonada. La feature degrada con `sin_datos`.
    """
    configurada = get_settings().eventos_ow_path.strip()
    if not configurada:
        return None
    ruta = Path(configurada)
    if not ruta.is_absolute():
        ruta = Path(__file__).resolve().parents[3] / configurada
    return ruta if ruta.exists() else None


def _parsear(ruta: Path) -> list[EventoOW]:
    import openpyxl

    libro = openpyxl.load_workbook(ruta, data_only=True, read_only=False)
    hoja = libro["Sheet1"]

    eventos: list[EventoOW] = []
    for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, values_only=True):
        campo = fila[_COL_CAMPO]
        inicio = fila[_COL_INICIO]
        if not campo or not isinstance(inicio, datetime):
            continue

        fin_bruto = fila[_COL_FIN]
        fin = (
            fin_bruto
            if (
                isinstance(fin_bruto, datetime)
                and _ANIO_MIN <= fin_bruto.year <= _ANIO_MAX
            )
            else None
        )

        tipo = ""
        for indice in _COLS_TIPO:
            valor = fila[indice]
            if valor and str(valor).strip():
                tipo = str(valor).strip()
                break

        eventos.append(
            EventoOW(
                campo=normalizar(str(campo)),
                pozo=str(fila[_COL_POZO] or "").strip(),
                tipo=tipo or "Evento",
                inicio=inicio,
                fin=fin,
            )
        )

    libro.close()
    return eventos


def cargar_eventos() -> list[EventoOW] | None:
    """Eventos del archivo, parseados UNA sola vez por proceso (A1).

    `None` = archivo ausente. Se reintenta en la siguiente llamada, por si
    aparece tras un despliegue.
    """
    global _cache_eventos
    if _cache_eventos is not None:
        return _cache_eventos

    ruta = ruta_eventos()
    if ruta is None:
        return None

    with _lock_carga:
        # Doble chequeo: otro hilo pudo cargarlo mientras esperábamos el lock.
        if _cache_eventos is None:
            logger.info("eventos_ow_parseando", ruta=str(ruta))
            _cache_eventos = _parsear(ruta)
            logger.info("eventos_ow_cargados", filas=len(_cache_eventos))
        return _cache_eventos


def reset_cache() -> None:
    """Vacía el singleton. Solo para tests."""
    global _cache_eventos
    with _lock_carga:
        _cache_eventos = None


class MantenimientosRepository:
    """Acceso al archivo de eventos. Sin BD."""

    def eventos(self) -> list[EventoOW] | None:
        return cargar_eventos()

    def campos_disponibles(self) -> set[str]:
        eventos = cargar_eventos() or []
        return {e["campo"] for e in eventos}
